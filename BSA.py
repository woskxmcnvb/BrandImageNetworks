import io 
from collections import OrderedDict

import pandas as pd
import numpy as np
from itertools import product

from openpyxl import load_workbook


from ExcelReportBuilder import ExcelReportBuilder
from GraphicModel import GraphicModel, PlotGraph, TupleToEdgeDict

from factor_analyzer import FactorAnalyzer

from pgmpy.estimators import HillClimbSearch, TreeSearch
from pgmpy.estimators import BicScore

import statsmodels.api as sm
from sklearn.feature_selection import SequentialFeatureSelector
from sklearn.feature_selection import RFECV
from sklearn.linear_model import LinearRegression

import utils

from definitions import *

SHEET_IMAGE = 'image_sets'
SHEET_DATA = 'data'
SHEET_MODEL = 'model_spec'


class ExtOrderedDict(OrderedDict):
    def top(self):
        return next(iter(self.items()))


def Standardize(df, across_all_dataframe=False): 
    if across_all_dataframe: 
        return df.sub(df.mean(axis=None)).div(df.stack().std())
    else:
        return df.sub(df.mean()).div(df.std())
    

def CalcFA(df: pd.DataFrame, n_factors) -> pd.DataFrame:
    fa = FactorAnalyzer(n_factors=n_factors, rotation='varimax').fit(df)
    return pd.DataFrame(fa.loadings_, index=df.columns, columns=['F{}'.format(i+1) for i in range(n_factors)])


def CalcBayesNetworkGraph(df: pd.DataFrame) -> io.BytesIO:
    hc = HillClimbSearch(df)
    model = hc.estimate(scoring_method=BicScore(df))
    stream = io.BytesIO()
    model.to_graphviz().draw(stream, format='png', prog='dot')
    return stream

def CalcBayesNetwork(df: pd.DataFrame):
    return HillClimbSearch(df).estimate(scoring_method=BicScore(df))

def CalcTree(df: pd.DataFrame, root_node: str): 
    return TreeSearch(df, root_node=root_node).estimate(estimator_type='chow-liu')


##################################################
############# First Order Dirvers ################
##################################################

def CalcLinearRegression(X: pd.DataFrame, y: pd.Series) -> pd.DataFrame: 
    model = sm.OLS(y, X).fit()
    return pd.read_html(io.StringIO(model.summary().tables[1].as_html()), index_col=0, header=0)[0]

def VariablesSelection(X: pd.DataFrame, y: pd.Series) -> pd.DataFrame: 
    model = LinearRegression()
    sfs = SequentialFeatureSelector(model).fit(X, y)
    rfe = RFECV(model).fit(X, y)
    return pd.concat([
        pd.Series(sfs.support_, index=X.columns), 
        pd.Series(rfe.support_, index=X.columns), 
        pd.Series(rfe.ranking_, index=X.columns)
    ], axis=1).set_axis(['Seq.Selector', 'Eliminator', 'Eliminator rank'], axis='columns')

def CalcLinRegWithVarSelect(X: pd.DataFrame, y: pd.Series, beta_tolerance=0.05) -> pd.DataFrame: 
    if y.name in X.columns:
        X = X.drop(y.name, axis='columns')
    result = pd.concat([CalcLinearRegression(X, y), VariablesSelection(X, y)], axis=1)
    result['Final Selection'] = (result['coef'] >= beta_tolerance) & (result['P>|t|'] <= 0.05) & result['Seq.Selector']
    return result

def SelectFirstOrderDrivers(X: pd.DataFrame, y: pd.DataFrame, beta_tolerance=0.05): 
    return pd.concat(
        [CalcLinRegWithVarSelect(X, y[col], beta_tolerance=beta_tolerance)['Final Selection'] for col in y.columns], 
        axis=1
    ).any(axis=1)

##################################################
##################################################


class ImageStructureAnalysis: 
    bn_ready: bool 
    fod_ready: bool

    image_sets: ExtOrderedDict
    expl_graph_model_sets: dict[str: GraphicModel]
    bsa_models: dict[str: GraphicModel]
    
    model_targets: list[str] = None
    mdf_based: bool = None

    clean_soe_data: pd.DataFrame = None 
    clean_targets_data: pd.DataFrame = None

    def __init__(self):
        self.bn_ready = False 
        self.fod_ready = False
        self.image_sets = ExtOrderedDict()
        self.expl_graph_model_sets = dict()
        self.bsa_models = dict()

    def __ReadImageSets(self): 
        data = pd.read_excel(self.data_file_name, sheet_name=SHEET_IMAGE, index_col=0)
        for set_ in data.columns: 
            if data[set_].notna().any(): 
                self.image_sets[set_] = data[set_].dropna().to_dict()
                self.expl_graph_model_sets[set_] = GraphicModel()

    def __ReadModelSpec(self):
        model_sheets = [sh for sh in load_workbook(self.data_file_name).sheetnames if sh.startswith(SHEET_MODEL)]
        self.bsa_models = dict()
        for sheet in model_sheets:
            name = sheet.split(SHEET_MODEL)[1]
            if name == '':
                name = str(len(self.bsa_models))
            spec = pd.read_excel(self.data_file_name, sheet_name=SHEET_MODEL)
            assert utils.AllElementsAreThere(EDGE_KEYS, spec.columns),\
                "Wrong model spec columns. There are: {}. Expected: {}".format(spec.columns, EDGE_KEYS)
            spec = spec[spec.notna().all(axis=1)]
            self.bsa_models[name] = GraphicModel(spec)


    def FullImageList(self):
        return self.image_sets.top()[1]
    
    def __SetModelTargets(self, mdf_based: bool, model_targets: str | list[str] | None) -> None: 
        self.mdf_based = mdf_based
        if self.mdf_based: 
            self.model_targets = FACTORS
            if model_targets:
                print("Warning: model_targets={} disregarded as mdf is used".format(model_targets))
            return
        if model_targets is None:
            raise ValueError("If no MDF, model_targets must be provided")
        if isinstance(model_targets, list):
            self.model_targets = model_targets
        elif isinstance(model_targets, str):
            self.model_targets = [model_targets]
        else:
            raise ValueError("Error: model_targets must be list[str] or str")
    
    def __ReadData(self):
        data = pd.read_excel(self.data_file_name, sheet_name=SHEET_DATA)

        cols_to_check = (list(MDF_RENAMER.keys()) if self.mdf_based else self.model_targets) +\
                        [SOE_COL_TEMPLATE.format(k) for k in self.FullImageList().keys()]
        missing_cols = utils.MissingColumn(data, cols_to_check)

        if missing_cols:
            if self.mdf_based:
                raise ValueError("Error: Missing columns in the data: {}".format(missing_cols))
            else:
                print("Warning: Check missing columns in the data: {}. It is ok only if they are constructs".format(missing_cols))
        
        na_filter = data[list(set(cols_to_check).difference(missing_cols))].notna().all(axis=1)

        soe_renamer = {SOE_COL_TEMPLATE.format(k): v for k, v in self.FullImageList().items()}
        self.clean_soe_data = data.loc[na_filter, soe_renamer.keys()].rename(columns=soe_renamer)
        if self.mdf_based:
            self.clean_targets_data = data.loc[na_filter, MDF_RENAMER.keys()].rename(columns=MDF_RENAMER)
        else:
            self.clean_targets_data = data.loc[na_filter, list(set(self.model_targets).difference(missing_cols))]

    def InitAndReadFile(self, file_name, mdf_based, model_targets, read_model_spec):
        self.data_file_name = file_name
        self.__SetModelTargets(mdf_based, model_targets)
        self.__ReadImageSets()
        self.__ReadData()

        if read_model_spec: 
            self.__ReadModelSpec()
        
        print("Info: Init and Read - Ok")
        return self
    
    def Inspect(self):
        print("MDF based-------- {}".format(self.mdf_based))
        print("Targets---------- {}".format(self.model_targets))
        print("Image sets--------{}".format(self.image_sets.keys()))
        print("Model specs-------{}".format(self.bsa_models.keys()))
        print("\nSOE data---------")
        if self.clean_soe_data:
            self.clean_soe_data.info(verbose=False, memory_usage=False)
        else:
            print('None')
        print("\nTargets data-----")
        if self.clean_targets_data:
            self.clean_targets_data.info(verbose=False, memory_usage=False)
        else:
            print('None')
    
    def CleanSOEandMDFData(self): 
        return pd.concat([self.clean_targets_data, self.clean_soe_data], axis=1)

    def ReportImageCrossCorrelations(self):
        corr_table = self.clean_soe_data.corr()
        # clean diagonal
        for r, c in product(range(len(corr_table)), range(len(corr_table))):
            if c == r:
                corr_table.iloc[r, c] = np.nan
        self.reporter.AddTable(corr_table, 'SOE cross-correlations', conditional_formatting=True)
        print("Info: SOE cross-correlations - Ok")

    def ReportFAs(self): 
        for i in range(3, len(self.FullImageList()) + 1):
            self.reporter.AddTable(CalcFA(self.clean_soe_data, i), 
                                   'factor_solutions', 
                                   conditional_formatting=True)
        print("Info: SOE FAs - Ok")
    
    def ReportSOEToEquityCorrelations(self):
        data = pd.concat([self.clean_targets_data, self.clean_soe_data], axis=1) 
        self.reporter.AddTable(
            data.corr().loc[self.clean_soe_data.columns, self.clean_targets_data.columns],
            'SOE to equity correlations', 
            conditional_formatting=True
        )
        print("Info: SOE to equity correlations - Ok")

    def ReportBayesNetworkGraphs(self): 
        for set_name, set_ in self.image_sets.items():
            self.reporter.AddImage(
                PlotGraph(CalcBayesNetwork(self.clean_soe_data[set_.values()]).to_graphviz()), 
                'Graphs_{}'.format(set_name)
            )
            # этот граф сохраняем для дальнейшего использования
            g = CalcBayesNetwork(self.clean_soe_data[set_.values()].mul(10).round()).to_graphviz()
            self.expl_graph_model_sets[set_name].AppendEdges(g)
            self.reporter.AddImage(PlotGraph(g), 'Graphs_{}'.format(set_name), "AA1")
        self.bn_ready = True
        print("Info: Bayes nets graphs - Ok")

    def ReportFirstOrderDrivers(self): 
        Y = Standardize(self.clean_targets_data, across_all_dataframe=False)
        XX = Standardize(self.clean_soe_data, across_all_dataframe=True)
        for set_name, set_ in self.image_sets.items():
            X = XX[set_.values()]
            for target_name, y in Y.items():
                fod = CalcLinRegWithVarSelect(X, y)
                # !!!!!!!!!!!!!!!!! ЗАГОЛОВОК TO REPORT
                self.reporter.AddTable(fod, 'FOD {}'.format(set_name))
                if target_name in self.model_targets:
                    self.expl_graph_model_sets[set_name].AppendEdges(
                        [TupleToEdgeDict((ix, target_name, EDGE_TYPE_PATH)) for ix, val in fod['Final Selection'].items() if val]
                    )
        self.fod_ready = True
        print("Info: First Order Drivers - Ok")

    def ReportBSAGraphs(self):
        if not self.bn_ready:
            self.ReportBayesNetworkGraphs()
        if not self.fod_ready:
            self.ReportFirstOrderDrivers() 

        for set_name, set_ in self.image_sets.items():
            if self.mdf_based:
                self.expl_graph_model_sets[set_name].AppendMDFGraph()
            self.expl_graph_model_sets[set_name].FitPLSPM(self.CleanSOEandMDFData()) 
            page_name = 'BSA_{}'.format(set_name)
            self.reporter.AddImage(
                PlotGraph(self.expl_graph_model_sets[set_name].Graph(add_pls_weights=False, exclude_mdf=True)), 
                page_name, "B1"
            )
            self.reporter.AddImage(
                PlotGraph(self.expl_graph_model_sets[set_name].Graph(add_pls_weights=True, exclude_mdf=False)), 
                page_name, "AA1"
            )
            self.reporter.AddTable(
                self.expl_graph_model_sets[set_name].model_spec.ToDF(),
                page_name, drop_index=True
            )

    def ReportTreeGraphs(self):
        for col in self.clean_soe_data: 
            self.reporter.AddImage(
                PlotGraph(CalcTree(self.clean_soe_data, col).to_graphviz()), 
                'TREE {}'.format(col)
            )
        print("Info: Tree Graphs - Ok")

    def EploratoryAnalisys(self,
                           file_name, 
                           mdf_based, 
                           model_targets=None,
                           report_correlations=True,
                           report_factor_solutions=False,
                           report_auto_graphs=False,
                           report_tree_graphs=False):
        # performs analysis, drops Excel report

        self.InitAndReadFile(file_name, mdf_based=mdf_based, model_targets=model_targets, read_model_spec=False)
        self.reporter = ExcelReportBuilder(file_name.split('.')[0] + '_exploratory_report.xlsx')
        
        if report_correlations:
            self.ReportImageCrossCorrelations()
            self.ReportSOEToEquityCorrelations()
        if report_factor_solutions:
            self.ReportFAs()
        if report_auto_graphs:
            self.ReportBSAGraphs()
        if report_tree_graphs:
            self.ReportTreeGraphs()

        self.reporter.SaveToFile()

    def BSAAnalysis(self, 
                    file_name, 
                    mdf_based, 
                    model_targets=None):
        
        self.InitAndReadFile(file_name, mdf_based=mdf_based, model_targets=model_targets, read_model_spec=True)
        for model in self.bsa_models.keys():
            self.__BSAReport(model)

    def __BSAReport(self, spec_name):
        reporter = ExcelReportBuilder("{}_bsa_report_{}.xlsx ".format(self.data_file_name.split('.')[0], spec_name))
        if self.mdf_based:
            self.bsa_models[spec_name].AppendMDFGraph()
        self.bsa_models[spec_name].FitPLSPM(self.CleanSOEandMDFData())

        reporter.AddTable(
            pd.concat(
                [self.bsa_models[spec_name].PathLenFromAllNodes(t) for t in self.model_targets], 
                axis=1
            ).set_axis(self.model_targets, axis='columns'), 
            "Importance"
        )

        reporter.AddImage(
            PlotGraph(self.bsa_models[spec_name].Graph(add_pls_weights=False, exclude_mdf=True)), 
            'GRAPHS'
        )
        reporter.AddImage(
            PlotGraph(self.bsa_models[spec_name].Graph(add_pls_weights=True, exclude_mdf=False)), 
            'GRAPHS',
            "AA1"
        )
        
        for node in self.bsa_models[spec_name].model_spec.Nodes():
            if (node in MDFS) or (node in POWER_PREMIUM) or (node in self.model_targets):
                continue
            reporter.AddImage(
                PlotGraph(self.bsa_models[spec_name].SubgraphFromNodes(node)), 
                node
        )

        reporter.SaveToFile()
        print("Info: model {} complete".format(spec_name))


    
    """def BSAReports(self, data_file_name):
        self.ReadDataFile(data_file_name)
        self.__ReadModelSpec()

        for model in self.bsa_models.keys():
            self.__OneBSAReport(model)"""

    """def __OneBSAReport(self, spec_name):
        reporter = ExcelReportBuilder("{}_bsa_report_{}.xlsx ".format(self.data_file_name.split('.')[0], spec_name))
        self.bsa_models[spec_name].AppendMDFGraph()
        self.bsa_models[spec_name].FitPLSPM(self.CleanSOEandMDFData())

        targets = POWER_PREMIUM + MDFS + FACTORS
        reporter.AddTable(
            pd.concat(
                [self.bsa_models[spec_name].PathLenFromAllNodes(t) for t in targets], 
                axis=1
            ).set_axis(targets, axis='columns'), 
            "importance"
        )

        reporter.AddImage(
            PlotGraph(self.bsa_models[spec_name].Graph(add_pls_weights=False, exclude_mdf=True)), 
            'TOTAL_GRAPHS'
        )
        reporter.AddImage(
            PlotGraph(self.bsa_models[spec_name].Graph(add_pls_weights=True, exclude_mdf=False)), 
            'TOTAL_GRAPHS',
            "AA1"
        )
        
        for node in self.bsa_models[spec_name].model_spec.Nodes():
            if (node in MDFS) or (node in POWER_PREMIUM):
                continue
            reporter.AddImage(
                PlotGraph(self.bsa_models[spec_name].SubgraphFromNodes(node)), 
                "GR_FROM_{}".format(node)
        )

        reporter.SaveToFile()"""






        
    
